import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font


# File Path
inputFile = r"C:\Users\arvindkumarpandit.a\Documents\PyCode\CrawlingData_Input.xlsx"
outputFile = r"C:\Users\arvindkumarpandit.a\Documents\PyCode\CrawlingData_Output.xlsx"
strSheetName = "CrawlingData"


#................Methods.....................................

def getManufactureName(parElement):    
    strMName = ''
    for pe in parElement:
        listFind=pe.find_all("div", {"class" : "head-container clearfix"})
        for sf in listFind:
            strElement=sf.find("a",{"class" : "brand solr-brand"})
            strMName=strElement.getText()
    return strMName      

def geUOM(parElement) :
    strUomName=''
    for pe in parElement:
        listFind=pe.find_all("div" , {"class" : "priceWrapper"})
        for sf in listFind:
            strElement=sf.find("span",{"class" : "gcprice-unit"})
            strUom=strElement.getText()
            if(strUom=='/ each'):
                strUomName = 'each 1'
            else:
                strUomName = 'pack'
    return strUomName
                    
def getModelName(parElement) :
    strModelName = ''
    for pe in parElement:
        listFind=pe.find_all("div", {"class" : "head-container clearfix"})
        for sf in listFind:
            strElement=sf.find("span" , {"itemprop" : "model"})
            strModelName=strElement.getText()
    return strModelName 

def getManufactureDescription(parElement) :
    strMDescription = ''
    for pe in parElement:
        listFind=pe.find_all("div", {"class" : "solr-breadcrumb-container"})
        for sf in listFind:
            strElement=sf.find("div" , {"class" : "bread-name last"})
            strMDescription=strElement.getText()
    return strMDescription

#................END OF METHODS..........................       

# Fetching the excel file, Hence please mention the proper excel file path and sheet
df = pd.read_excel(inputFile, sheet_name=strSheetName)
wb = openpyxl.load_workbook(inputFile)
sheet1 = wb.get_sheet_by_name(strSheetName)
wsheet = wb.active

# Selecting the columns which are required to be compared
sdf = df[['Sl_No', 'Master_Number', 'Master_Description','Supplier_Name']]

# Finding number of rows and columns
totalRows = len(sdf.axes[0])
totalcols = len(df.axes[1])
colIndex = totalcols + 1

#Columns Setting
sheet1.cell(row=1, column=colIndex).value = 'Manufacture_Name'
sheet1.cell(row=1, column=colIndex).font = Font(bold=True)
wsheet.column_dimensions['E'].width = 20
sheet1.cell(row=1, column=colIndex + 1).value = 'Manufacture_Part_Number'
sheet1.cell(row=1, column=colIndex + 1).font = Font(bold=True)
wsheet.column_dimensions['F'].width = 20
sheet1.cell(row=1, column=colIndex + 2).value = 'Manufacture_Description'
sheet1.cell(row=1, column=colIndex + 2).font = Font(bold=True)
wsheet.column_dimensions['G'].width = 20
sheet1.cell(row=1, column=colIndex + 3).value = 'UOM'
sheet1.cell(row=1, column=colIndex + 3).font = Font(bold=True)
wsheet.column_dimensions['H'].width = 20
sheet1.cell(row=1, column=colIndex + 4).value = 'Packsize'
sheet1.cell(row=1, column=colIndex + 4).font = Font(bold=True)
wsheet.column_dimensions['I'].width = 20

# Row Comparison
for i in range(0, totalRows):

    cell = sheet1.cell(row=i + 2, column= 2)
    url="https://www.grainger.com/product/"+str(cell.value)+"?searchQuery="+str(cell.value).lower()+"&searchBar=true"
    #url="https://www.grainger.com/product/3GUH4?searchBar=true&searchQuery=3GUH4"
    strRequest=requests.get(url)
    strSoup = BeautifulSoup(strRequest.text)
    allElements=strSoup.find_all("div" , {"id" : "body"})

    if cell.value is not None:
        if not str(cell.value).isalpha():
            if not str(cell.value).isdigit():
                if(str(cell.value).isalnum()):
                    sheet1.cell(row=i + 2, column=colIndex).value = getManufactureName(allElements)
                    sheet1.cell(row=i + 2, column=colIndex + 1).value = getModelName(allElements)
                    sheet1.cell(row=i + 2, column=colIndex + 2).value = getManufactureDescription(allElements)
                    sheet1.cell(row=i + 2, column=colIndex + 3).value = geUOM(allElements)
                    sheet1.cell(row=i + 2, column=colIndex + 4).value = 'Pack Size ...'
        

# Saving excel output file, provide path for save the output file
wb.save(outputFile)
