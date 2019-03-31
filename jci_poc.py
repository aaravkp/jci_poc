from tkinter import *
from tkinter import filedialog
import pandas as pd
import openpyxl
import os

window = Tk()
window.title("JCI Window")
window.geometry("600x550")


#Functions
def cmdLoad():
    strfilename= filedialog.askopenfilename()
    txtFilename.set(strfilename)
    df = pd.read_excel(strfilename, sheetname='Master')
    for item in df.columns:
        list1.insert(END, item)


def cmdMove():
    selected = list1.curselection()
    if selected:
        for item in selected:
            list2.insert(END, list1.get(item))

def cmdRemove():
    items = list2.curselection()
    pos = 0
    for i in items:
        idx = int(i) - pos
        list2.delete(idx, idx)
        pos = pos + 1

def cmdSubmit():
    df = pd.read_excel(txtFilename.get(), sheetname='Master')
    wb = openpyxl.load_workbook(txtFilename.get())
    sheet1 = wb.get_sheet_by_name('Master')

    for item in range(list2.size()):
        lstvalue=list2.get(item)
        for col in df.columns:
            if (lstvalue == col):
                for row in range(1, sheet1.max_row + 1):
                    columnIndex = int(df.columns.get_loc(col)) + 1
                    cell = sheet1.cell(row=row, column= columnIndex)
                    if cell.value is None:
                        cell.value = "Updated"


    wb.save(os.path.dirname(txtFilename.get()) + '/OutputFile.xlsx')
    lmsg = Label(window, text="File data is updated successfully.", fg="green")
    lmsg.grid(row=5, column=1)


#Lables

l1 = Label(window, text="Load File:")
l1.grid(row=1,column=0)

l1 = Label(window, text="Select Columns:")
l1.grid(row=2,column=0)

l1 = Label(window, text="Selected Columns:")
l1.grid(row=4,column=0)

#ListBox

list1= Listbox(window, selectmode='multiple', height=6, width=35)
list1.grid(row=2, column=1)

list2= Listbox(window, height=6, width=35)
list2.grid(row =4, column=1)

#Entries
txtFilename = StringVar()
e1 = Entry(window, textvariable=txtFilename)
e1.grid(row=1, column=1)
#Define Button
b1= Button(window, text="Load", command=cmdLoad, width=12, fg="blue")
b1.grid(row=1, column=2, pady=20, padx=8)

b2= Button(window, text="Move>>", command=cmdMove, width=12, fg="blue")
b2.grid(row=3, column=1, pady=20, padx=8)

b3= Button(window, text="Submit", width=12, command=cmdSubmit, fg="blue")
b3.grid(row=5, column=1, pady=20, padx=8)

b4= Button(window, text="Close", width=12, command=quit, fg="red")
b4.grid(row=5, column=2, pady=20, padx=8)

b5= Button(window, text="Remove", command=cmdRemove, width=12, fg="red")
b5.grid(row=4, column=2, pady=20, padx=8)

window.mainloop()