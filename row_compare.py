import pandas as pd
import openpyxl

# File Path
inputFile = '/Users/arvindkumar/Documents/JCI/Input/Book1.xlsx'
outputFile = '/Users/arvindkumar/Documents/JCI/Output/OutputFile.xlsx'

sheetName = 'Sheet1'

# Fetching the excel file, Hence please mention the proper excel file path and sheet
df = pd.read_excel(inputFile, sheetname=sheetName)
wb = openpyxl.load_workbook(inputFile)
sheet1 = wb.get_sheet_by_name(sheetName)

# Selecting the columns which are required to be compared
sdf = df[['Part name', 'Product Name', 'Part Cost','Part Number']]

# Finding number of rows and columns
totalRows = len(sdf.axes[0])
totalcols = len(df.axes[1])

colIndex = totalcols + 1

sheet1.cell(row=1, column=colIndex).value = 'Match Found'
matchCount = 1
for i in range(0, totalRows):

    for j in range(i+1, totalRows):

        if list(sdf.loc[i,:]) == list(sdf.loc[j,:]):
            if sheet1.cell(row=i + 2, column=colIndex).value is None or sheet1.cell(row=j+2, column=colIndex).value is None:
                if i == 0:
                    sheet1.cell(row=i + 2, column=colIndex).value = matchCount
                    sheet1.cell(row=j + 2, column=colIndex).value = matchCount
                    matchCount = sheet1.cell(row=j + 2, column=colIndex).value
                else:
                    sheet1.cell(row=i + 2, column=colIndex).value = matchCount + 1
                    sheet1.cell(row=j + 2, column=colIndex).value = matchCount + 1
                    matchCount = sheet1.cell(row=j + 2, column=colIndex).value

# Saving excel output file, provide path for save the output file
wb.save(outputFile)
